from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(r"D:\Code\阿里国际站")
DESKTOP_DIR = Path.home() / "Desktop" / "Alibaba_Store_Audit_2026-04-04"
TODAY = "2026-04-04"
FONT_PATH = Path(r"C:\Windows\Fonts\simhei.ttf")


def load_json(prefix: str) -> dict:
    path = next(DESKTOP_DIR.glob(f"{prefix}_*.json"))
    return json.loads(path.read_text(encoding="utf-8"))


def load_csv_rows(prefix: str) -> list[dict]:
    path = next(DESKTOP_DIR.glob(f"{prefix}_*.csv"))
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_text(name: str) -> str:
    return (DESKTOP_DIR / name).read_text(encoding="utf-8")


def ensure_font() -> None:
    if "SimHei" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("SimHei", str(FONT_PATH)))


def workbook_base() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "封面"
    ws["A1"] = "阿里国际站经营交付"
    ws["A2"] = f"生成日期：{TODAY}"
    ws["A3"] = "说明：Excel 为正式工作底稿，Markdown/JSON/CSV 为附录。"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"].font = Font(size=11)
    ws["A3"].font = Font(size=11, color="666666")
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 50
    return wb


def style_sheet(ws, freeze: str = "A2", autofilter: bool = True) -> None:
    ws.freeze_panes = freeze
    if autofilter and ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, 40))
        ws.column_dimensions[letter].width = max_len


def add_table_sheet(wb: Workbook, title: str, rows: list[dict], preferred: list[str] | None = None) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws["A1"] = "暂无数据"
        return
    headers = preferred or list(rows[0].keys())
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row.get(header))
    style_sheet(ws)


def normalize_module_status(data: dict, account: str) -> list[dict]:
    rows = []
    for module, status in data.get("moduleStatus", {}).items():
        if isinstance(status, dict):
            parts = []
            for key in ["status", "source", "verification", "note", "reason"]:
                value = status.get(key)
                if value:
                    parts.append(f"{key}: {value}")
            normalized = " | ".join(parts) if parts else json.dumps(status, ensure_ascii=False)
        else:
            normalized = status
        rows.append(
            {
                "账号": account.upper(),
                "模块": module,
                "状态": normalized,
            }
        )
    return rows


def build_metrics_rows(wika: dict, xd: dict) -> list[dict]:
    return [
        {
            "指标名称": "产品总数",
            "粒度": "产品",
            "WIKA 是否可得": "是",
            "XD 是否可得": "是",
            "数据源类型": "官方 API",
            "是否生产无状态": "是",
            "当前状态": "已验证",
            "接口/页面/导出来源说明": "products/list 官方主数据",
            "后续动作": "继续补产品表现层",
        },
        {
            "指标名称": "产品标题",
            "粒度": "产品",
            "WIKA 是否可得": "是",
            "XD 是否可得": "是",
            "数据源类型": "官方 API",
            "是否生产无状态": "是",
            "当前状态": "已验证",
            "接口/页面/导出来源说明": "products/list -> subject",
            "后续动作": "接入统一产品视图",
        },
        {
            "指标名称": "产品分组/系列",
            "粒度": "产品",
            "WIKA 是否可得": "是",
            "XD 是否可得": "是",
            "数据源类型": "官方 API",
            "是否生产无状态": "是",
            "当前状态": "已验证",
            "接口/页面/导出来源说明": "products/list -> group_name",
            "后续动作": "纳入系列化诊断",
        },
        {
            "指标名称": "上架时间/最近修改时间",
            "粒度": "产品",
            "WIKA 是否可得": "是",
            "XD 是否可得": "是",
            "数据源类型": "官方 API",
            "是否生产无状态": "是",
            "当前状态": "已验证",
            "接口/页面/导出来源说明": "products/list -> gmt_create / gmt_modified",
            "后续动作": "形成产品老化分析",
        },
        {
            "指标名称": "产品曝光/点击/CTR",
            "粒度": "产品",
            "WIKA 是否可得": "部分",
            "XD 是否可得": "否",
            "数据源类型": "页面态",
            "是否生产无状态": "否",
            "当前状态": "WIKA 已验证 / XD 暂无来源",
            "接口/页面/导出来源说明": "WIKA 仅本地页面态 performance30d；XD 未接通",
            "后续动作": "优先识别 XD 可复用生产源，否则维持不可得",
        },
        {
            "指标名称": "店铺 UV/PV",
            "粒度": "店铺",
            "WIKA 是否可得": "是",
            "XD 是否可得": "否",
            "数据源类型": "页面态",
            "是否生产无状态": "否",
            "当前状态": "WIKA 已验证 / XD 暂无来源",
            "接口/页面/导出来源说明": "WIKA overview 页面态；XD overview 未接通",
            "后续动作": "识别 XD overview 生产无状态源",
        },
        {
            "指标名称": "询盘数",
            "粒度": "店铺",
            "WIKA 是否可得": "部分",
            "XD 是否可得": "否",
            "数据源类型": "页面态",
            "是否生产无状态": "否",
            "当前状态": "WIKA 页面态 / XD 权限或来源未定",
            "接口/页面/导出来源说明": "WIKA overview 页面态可见；XD 未接通",
            "后续动作": "继续官方/无状态来源识别",
        },
        {
            "指标名称": "国家/地区来源",
            "粒度": "店铺",
            "WIKA 是否可得": "是",
            "XD 是否可得": "否",
            "数据源类型": "页面态",
            "是否生产无状态": "否",
            "当前状态": "WIKA 已验证 / XD 暂无来源",
            "接口/页面/导出来源说明": "WIKA marketProfile 页面态；XD overview 未接通",
            "后续动作": "作为 XD overview 核心识别目标",
        },
        {
            "指标名称": "订单量",
            "粒度": "订单",
            "WIKA 是否可得": "是",
            "XD 是否可得": "是",
            "数据源类型": "官方 API",
            "是否生产无状态": "是",
            "当前状态": "已验证",
            "接口/页面/导出来源说明": "orders/list -> totalCount",
            "后续动作": "补 orders 汇总和趋势层",
        },
        {
            "指标名称": "订单明细",
            "粒度": "订单",
            "WIKA 是否可得": "是",
            "XD 是否可得": "是",
            "数据源类型": "官方 API",
            "是否生产无状态": "是",
            "当前状态": "已验证",
            "接口/页面/导出来源说明": "orders/detail -> alibaba.seller.order.get",
            "后续动作": "扩字段并沉淀统一结构",
        },
        {
            "指标名称": "订单金额/状态/发货方式",
            "粒度": "订单",
            "WIKA 是否可得": "是",
            "XD 是否可得": "部分",
            "数据源类型": "WIKA 页面态 + 官方 detail",
            "是否生产无状态": "部分",
            "当前状态": "WIKA 已验证 / XD 待扩展",
            "接口/页面/导出来源说明": "WIKA 有本地分析层；XD 目前仅 detail 样本字段",
            "后续动作": "先补 XD orders 最小汇总",
        },
        {
            "指标名称": "询盘/RFQ 列表与来源",
            "粒度": "询盘",
            "WIKA 是否可得": "否",
            "XD 是否可得": "否",
            "数据源类型": "官方候选存在",
            "是否生产无状态": "未验证",
            "当前状态": "当前未识别到可用入口",
            "接口/页面/导出来源说明": "RFQ / inquiry 公开能力存在，但尚未证明可用于当前生产认证体系",
            "后续动作": "继续只做官方接口可行性识别，不进入路由开发",
        },
        {
            "指标名称": "消息列表/回复记录",
            "粒度": "询盘",
            "WIKA 是否可得": "否",
            "XD 是否可得": "否",
            "数据源类型": "官方候选弱",
            "是否生产无状态": "未验证",
            "当前状态": "当前未识别到可用入口",
            "接口/页面/导出来源说明": "仅发现翻译设置/发送类线索，未发现已验证消息读取接口",
            "后续动作": "保持停在识别阶段",
        },
        {
            "指标名称": "客户通客户画像",
            "粒度": "客户",
            "WIKA 是否可得": "否",
            "XD 是否可得": "否",
            "数据源类型": "官方 API",
            "是否生产无状态": "未验证",
            "当前状态": "官方存在但权限阻塞",
            "接口/页面/导出来源说明": "customer.batch.get / customer.get / customer.note.* 文档存在，但标注聚石塔内调用",
            "后续动作": "先确认权限/上下文，再决定是否进入路由开发",
        },
    ]


def build_summary_rows(wika: dict, xd: dict) -> list[dict]:
    wika_products = wika["products"]["fullCatalogSummary"]
    xd_products = xd["products"]["fullCatalogSummary"]
    return [
        {"账号": "WIKA", "产品总数": wika_products["total_item"], "近30天更新": wika_products["updated_in_30_days"], "180天未更新": wika_products["stale_over_180_days"], "订单总数": wika["orders"]["listSnapshot"]["totalCount"], "运行状态": wika["runtimeStatus"]["runtimeLoadedFrom"]},
        {"账号": "XD", "产品总数": xd_products["total_item"], "近30天更新": xd_products["updated_in_30_days"], "180天未更新": xd_products["stale_over_180_days"], "订单总数": xd["orders"]["listSnapshot"]["totalCount"], "运行状态": xd["runtimeStatus"]["runtimeLoadedFrom"]},
    ]


def top_group_rows(account: str, data: dict) -> list[dict]:
    rows = []
    for item in data["products"]["fullCatalogSummary"]["top_groups"]:
        rows.append({"账号": account, "分组": item["group_name"], "产品数": item["count"]})
    return rows


def order_status_rows(account: str, data: dict) -> list[dict]:
    rows = []
    dist = data.get("orders", {}).get("orderStatusDistribution") or []
    for item in dist:
        rows.append({"账号": account, "订单状态": item.get("status_name"), "数量": item.get("count")})
    return rows


def diagnostics_rows(account: str, data: dict) -> list[dict]:
    rows = []
    diagnostics = data.get("diagnostics", {})
    mapping = {"factual": "真实数据结论", "inference": "基于结构的推断", "unknown": "因数据缺失暂不能判断"}
    for key, label in mapping.items():
        for item in diagnostics.get(key, []):
            rows.append({"账号": account, "结论类型": label, "内容": item})
    return rows


def sample_product_rows(account: str, data: dict) -> list[dict]:
    rows = []
    for item in data.get("products", {}).get("sampleProducts", [])[:15]:
        rows.append(
            {
                "账号": account,
                "product_id": item.get("product_id"),
                "标题": item.get("subject"),
                "分组": item.get("group_name"),
                "状态": item.get("status"),
                "最近修改时间": item.get("gmt_modified"),
                "详情链接": item.get("pc_detail_url"),
            }
        )
    return rows


def sample_order_rows(account: str, data: dict) -> list[dict]:
    rows = []
    for item in data.get("orders", {}).get("listSnapshot", {}).get("sampleItems", [])[:20]:
        create_date = item.get("create_date") or {}
        modify_date = item.get("modify_date") or {}
        rows.append(
            {
                "账号": account,
                "trade_id": item.get("trade_id"),
                "创建时间": create_date.get("format_date") or create_date.get("timestamp"),
                "修改时间": modify_date.get("format_date") or modify_date.get("timestamp"),
            }
        )
    return rows


def issue_rows(wika: dict, xd: dict) -> list[dict]:
    return [
        {"账号": "WIKA", "问题级别": "高", "类型": "真实数据结论", "问题": "美国访客占比高，市场集中度风险明显。", "来源": "WIKA marketProfile 页面态真实数据"},
        {"账号": "WIKA", "问题级别": "中", "类型": "真实数据结论", "问题": "180 天未更新产品数量高，存在内容老化。", "来源": "WIKA 官方 products 主数据"},
        {"账号": "XD", "问题级别": "高", "类型": "真实数据结论", "问题": "overview 未接通，无法判断真实流量来源与市场结构。", "来源": "XD 模块状态"},
        {"账号": "XD", "问题级别": "高", "类型": "基于结构的推断", "问题": "目录过宽，非包装主线稀释了店铺定位。", "来源": "XD 产品分组结构 + 行业对标"},
        {"账号": "XD", "问题级别": "中", "类型": "真实数据结论", "问题": "orders 仅完成最小 list/detail，尚无完整汇总与趋势。", "来源": "XD orders 模块状态"},
    ]


def suggestion_rows() -> list[dict]:
    return [
        {"账号": "WIKA", "周期": "短期(1-2周)", "优先级": "高", "建议类型": "官方运营知识对应建议", "建议": "围绕包装套装、镜布、清洁套件补齐 MOQ、打样、交期、logo 工艺首屏表达。"},
        {"账号": "WIKA", "周期": "中期(1-2月)", "优先级": "中", "建议类型": "基于结构的推断", "建议": "围绕美国以外高潜市场补国家级产品组合与详情页文案。"},
        {"账号": "XD", "周期": "短期(1-2周)", "优先级": "高", "建议类型": "真实数据结论", "建议": "先收口目录，把包装主线 case/pouch/cloth/cleaner 提到前台。"},
        {"账号": "XD", "周期": "中期(1-2月)", "优先级": "高", "建议类型": "真实数据结论", "建议": "在已验证 orders 官方 list/detail 之上补最小汇总与趋势层。"},
        {"账号": "WIKA/XD", "周期": "长期", "优先级": "中", "建议类型": "官方运营知识对应建议", "建议": "将 MOQ、打样、交期、材质、工厂能力、环保与证书做成固定化 B2B 采购表达模板。"},
    ]


def add_summary_chart(ws, start_row: int, rows: list[dict]) -> None:
    for idx, row in enumerate(rows, start=start_row):
        ws.cell(idx, 1, row["账号"])
        ws.cell(idx, 2, row["产品总数"])
        ws.cell(idx, 3, row["订单总数"])
    for col in range(1, 4):
        ws.cell(start_row - 1, col, ["账号", "产品总数", "订单总数"][col - 1]).font = Font(bold=True)
    chart = BarChart()
    chart.title = "WIKA / XD 产品与订单规模对比"
    chart.y_axis.title = "数量"
    data = Reference(ws, min_col=2, max_col=3, min_row=start_row - 1, max_row=start_row + len(rows) - 1)
    cats = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + len(rows) - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, "H4")


def create_main_workbook(wika: dict, xd: dict, metrics_rows: list[dict]) -> Path:
    wb = workbook_base()
    summary_rows = build_summary_rows(wika, xd)

    cover = wb["封面"]
    cover["A5"] = "执行摘要"
    cover["A5"].font = Font(bold=True, size=14)
    summary_text = [
        "WIKA 与 XD 都已具备生产 auth/bootstrap 闭环。",
        "WIKA 当前在产品、订单、市场分析层成熟度高于 XD。",
        "XD 已完成产品主数据与最小 orders 官方 list/detail，overview 与 inquiries 仍未接通。",
        "本工作簿优先面向经营管理，原始 JSON/CSV 作为附录使用。",
    ]
    for idx, line in enumerate(summary_text, start=6):
        cover[f"A{idx}"] = f"• {line}"
    add_summary_chart(cover, 13, summary_rows)

    add_table_sheet(wb, "摘要总表", summary_rows)
    add_table_sheet(wb, "指标覆盖矩阵", metrics_rows)
    add_table_sheet(wb, "WIKA_分组", top_group_rows("WIKA", wika))
    add_table_sheet(wb, "XD_分组", top_group_rows("XD", xd))
    add_table_sheet(wb, "WIKA_产品样本", sample_product_rows("WIKA", wika))
    add_table_sheet(wb, "XD_产品样本", sample_product_rows("XD", xd))
    add_table_sheet(wb, "WIKA_订单样本", sample_order_rows("WIKA", wika))
    add_table_sheet(wb, "XD_订单样本", sample_order_rows("XD", xd))
    add_table_sheet(wb, "订单状态", order_status_rows("WIKA", wika) + order_status_rows("XD", xd))
    add_table_sheet(wb, "诊断结论", diagnostics_rows("WIKA", wika) + diagnostics_rows("XD", xd))
    add_table_sheet(wb, "问题清单", issue_rows(wika, xd))
    add_table_sheet(wb, "建议清单", suggestion_rows())

    output = DESKTOP_DIR / f"WIKA_XD_运营总报告_{TODAY}.xlsx"
    wb.save(output)
    return output


def create_module_status_workbook(wika: dict, xd: dict, metrics_rows: list[dict]) -> Path:
    wb = workbook_base()
    wb["封面"]["A1"] = "WIKA / XD 模块状态与数据覆盖"
    add_table_sheet(wb, "模块状态", normalize_module_status(wika, "wika") + normalize_module_status(xd, "xd"))
    add_table_sheet(wb, "数据覆盖", metrics_rows)
    output = DESKTOP_DIR / f"WIKA_XD_模块状态与数据覆盖_{TODAY}.xlsx"
    wb.save(output)
    return output


def create_metrics_matrix_artifacts(metrics_rows: list[dict]) -> tuple[Path, Path]:
    csv_path = DESKTOP_DIR / f"WIKA_XD_指标覆盖矩阵_{TODAY}.csv"
    xlsx_path = DESKTOP_DIR / f"WIKA_XD_指标覆盖矩阵_{TODAY}.xlsx"

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(metrics_rows[0].keys()))
        writer.writeheader()
        writer.writerows(metrics_rows)

    wb = workbook_base()
    wb["封面"]["A1"] = "WIKA / XD 指标覆盖矩阵"
    add_table_sheet(wb, "矩阵", metrics_rows)
    wb.save(xlsx_path)
    return csv_path, xlsx_path


def create_industry_workbook() -> Path:
    rows = load_csv_rows("行业领头企业对比表")
    wb = workbook_base()
    wb["封面"]["A1"] = "行业领头企业对比与学习建议"
    add_table_sheet(wb, "对比表", rows)
    output = DESKTOP_DIR / f"行业领头企业对比与学习建议_{TODAY}.xlsx"
    wb.save(output)
    return output


def pdf_styles():
    ensure_font()
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="CnTitle", parent=styles["Title"], fontName="SimHei", fontSize=20, leading=26))
    styles.add(ParagraphStyle(name="CnHeading1", parent=styles["Heading1"], fontName="SimHei", fontSize=15, leading=20))
    styles.add(ParagraphStyle(name="CnHeading2", parent=styles["Heading2"], fontName="SimHei", fontSize=12, leading=16))
    styles.add(ParagraphStyle(name="CnBody", parent=styles["BodyText"], fontName="SimHei", fontSize=9.5, leading=14))
    return styles


def table_style() -> TableStyle:
    return TableStyle(
        [
            ("FONTNAME", (0, 0), (-1, -1), "SimHei"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9E2F3")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FBFF")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]
    )


def create_main_pdf(wika: dict, xd: dict, metrics_rows: list[dict]) -> Path:
    styles = pdf_styles()
    output = DESKTOP_DIR / f"WIKA_XD_运营总报告_{TODAY}.pdf"
    doc = SimpleDocTemplate(str(output), pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    story = []
    story.append(Paragraph("WIKA + XD 运营总报告", styles["CnTitle"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"生成日期：{TODAY}", styles["CnBody"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph("执行摘要", styles["CnHeading1"]))
    for line in [
        "WIKA 与 XD 当前都已完成生产授权、runtime 持久化和 bootstrap 冷启动恢复。",
        "WIKA 在产品、订单、市场分析层成熟度明显高于 XD；XD 目前完成的是产品主数据与 orders 最小官方链路。",
        "XD overview、XD inquiries/messages/customers 仍无已验证生产无状态数据源，当前不能伪诊断。",
        "本报告区分真实数据结论、结构推断、官方运营知识建议与暂不能判断项。",
    ]:
        story.append(Paragraph(f"• {line}", styles["CnBody"]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("店铺关键摘要", styles["CnHeading1"]))
    summary_rows = build_summary_rows(wika, xd)
    summary_table = [["账号", "产品总数", "近30天更新", "180天未更新", "订单总数", "运行状态"]]
    for row in summary_rows:
        summary_table.append([row["账号"], row["产品总数"], row["近30天更新"], row["180天未更新"], row["订单总数"], row["运行状态"]])
    table = Table(summary_table, colWidths=[22 * mm, 22 * mm, 22 * mm, 24 * mm, 22 * mm, 48 * mm])
    table.setStyle(table_style())
    story.append(table)
    story.append(Spacer(1, 8))

    story.append(Paragraph("模块完成情况", styles["CnHeading1"]))
    module_rows = normalize_module_status(wika, "wika") + normalize_module_status(xd, "xd")
    module_table = [["账号", "模块", "状态"]] + [[r["账号"], r["模块"], r["状态"]] for r in module_rows]
    table = Table(module_table, colWidths=[18 * mm, 40 * mm, 100 * mm])
    table.setStyle(table_style())
    story.append(table)
    story.append(PageBreak())

    story.append(Paragraph("经营关键指标覆盖矩阵（节选）", styles["CnHeading1"]))
    cut_rows = metrics_rows[:10]
    metrics_table = [["指标名称", "粒度", "WIKA", "XD", "状态", "后续动作"]]
    for row in cut_rows:
        metrics_table.append([row["指标名称"], row["粒度"], row["WIKA 是否可得"], row["XD 是否可得"], row["当前状态"], row["后续动作"]])
    table = Table(metrics_table, colWidths=[30 * mm, 16 * mm, 16 * mm, 16 * mm, 42 * mm, 50 * mm])
    table.setStyle(table_style())
    story.append(table)
    story.append(Spacer(1, 8))

    story.append(Paragraph("当前经营关键数据缺口", styles["CnHeading1"]))
    for line in [
        "XD 店铺级 UV/PV、国家来源、流量来源目前无已验证生产无状态来源。",
        "WIKA 产品表现层与订单分析层仍依赖本地页面态，不适合直接复制为 XD 生产路径。",
        "WIKA / XD inquiries/messages/customers 当前没有已验证生产链路。",
    ]:
        story.append(Paragraph(f"• {line}", styles["CnBody"]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("核心诊断结论", styles["CnHeading1"]))
    for account, data in [("WIKA", wika), ("XD", xd)]:
        story.append(Paragraph(account, styles["CnHeading2"]))
        for section, label in [("factual", "真实数据结论"), ("inference", "基于结构的推断"), ("unknown", "暂不能判断")]:
            values = data.get("diagnostics", {}).get(section, [])
            if values:
                story.append(Paragraph(label, styles["CnBody"]))
                for item in values:
                    story.append(Paragraph(f"• {item}", styles["CnBody"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph("建议优先级", styles["CnHeading1"]))
    for item in suggestion_rows():
        story.append(
            Paragraph(
                f"• [{item['账号']}] {item['周期']} / {item['优先级']} / {item['建议类型']}：{item['建议']}",
                styles["CnBody"],
            )
        )
    doc.build(story)
    return output


def create_industry_pdf() -> Path:
    styles = pdf_styles()
    output = DESKTOP_DIR / f"行业领头企业分析报告_{TODAY}.pdf"
    doc = SimpleDocTemplate(str(output), pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    content = load_text(f"行业领头企业分析报告_{TODAY}.md")
    story = [Paragraph("行业领头企业分析报告", styles["CnTitle"]), Spacer(1, 6)]
    for line in content.splitlines():
        if not line.strip():
            story.append(Spacer(1, 4))
        elif line.startswith("# "):
            story.append(Paragraph(line[2:], styles["CnHeading1"]))
        elif line.startswith("## "):
            story.append(Paragraph(line[3:], styles["CnHeading2"]))
        elif line.startswith("- "):
            story.append(Paragraph(f"• {line[2:]}", styles["CnBody"]))
        else:
            story.append(Paragraph(line, styles["CnBody"]))
    doc.build(story)
    return output


def safe_generate(factory, expected_path: Path) -> Path:
    try:
        return factory()
    except PermissionError:
        if expected_path.exists():
            return expected_path
        raise


def update_readme(files: dict[str, Path]) -> Path:
    path = DESKTOP_DIR / "README_交付说明.md"
    content = f"""# 交付说明

生成日期：{TODAY}

## 主交付
- Excel 主报告：`{files['main_xlsx'].name}`
- Excel 模块状态与覆盖：`{files['module_xlsx'].name}`
- Excel 指标覆盖矩阵：`{files['matrix_xlsx'].name}`
- Excel 行业对标：`{files['industry_xlsx'].name}`
- PDF 主报告：`{files['main_pdf'].name}`
- PDF 行业报告：`{files['industry_pdf'].name}`

## 附录与原始数据
- `WIKA_完整数据_{TODAY}.json`
- `XD_完整数据_{TODAY}.json`
- `WIKA_XD_对比汇总_{TODAY}.csv`
- `行业领头企业对比表_{TODAY}.csv`
- `WIKA_XD_运营总报告_{TODAY}.md`

## 说明
- WIKA products 与最小 orders 官方路由已线上验证。
- XD products 与最小 orders 官方路由已线上验证。
- WIKA 的市场、产品表现、订单分析层仍包含本地页面态已验证来源。
- XD overview 与 inquiries/messages/customers 当前仍未接通，不在主报告中伪造为空数据。
"""
    path.write_text(content, encoding="utf-8")
    return path


def main() -> None:
    wika = load_json("WIKA")
    xd = load_json("XD")
    metrics_rows = build_metrics_rows(wika, xd)

    matrix_csv, matrix_xlsx = create_metrics_matrix_artifacts(metrics_rows)
    main_xlsx = create_main_workbook(wika, xd, metrics_rows)
    module_xlsx = create_module_status_workbook(wika, xd, metrics_rows)
    industry_xlsx = create_industry_workbook()
    main_pdf = safe_generate(lambda: create_main_pdf(wika, xd, metrics_rows), DESKTOP_DIR / f"WIKA_XD_运营总报告_{TODAY}.pdf")
    industry_pdf = safe_generate(create_industry_pdf, DESKTOP_DIR / f"行业领头企业分析报告_{TODAY}.pdf")
    readme = update_readme(
        {
            "matrix_csv": matrix_csv,
            "matrix_xlsx": matrix_xlsx,
            "main_xlsx": main_xlsx,
            "module_xlsx": module_xlsx,
            "industry_xlsx": industry_xlsx,
            "main_pdf": main_pdf,
            "industry_pdf": industry_pdf,
        }
    )

    print("generated:")
    for path in [matrix_csv, matrix_xlsx, main_xlsx, module_xlsx, industry_xlsx, main_pdf, industry_pdf, readme]:
        print(path)


if __name__ == "__main__":
    main()
